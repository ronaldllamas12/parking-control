"""API de administración financiera."""

from __future__ import annotations

import io
from datetime import datetime

from app import schemas
from app.database import get_db
from app.security import role_required
from app.services import finanzas_service
from app.services.telegram_service import enviar_notificacion_telegram
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session

router = APIRouter(prefix="/finanzas", tags=["finanzas"])


def _centavos_a_cop(centavos: int) -> str:
    return f"${centavos / 100:,.0f}"


# ── Config ────────────────────────────────────────────────────────────────────

@router.get("/config", response_model=schemas.ConfigFinancieraOut)
def get_config(
    current_user=Depends(role_required(["admin"])),
    db: Session = Depends(get_db),
):
    return finanzas_service.get_or_create_config(db, current_user.conjunto_id)


@router.put("/config", response_model=schemas.ConfigFinancieraOut)
def put_config(
    payload: schemas.ConfigFinancieraUpdate,
    current_user=Depends(role_required(["admin"])),
    db: Session = Depends(get_db),
):
    return finanzas_service.update_config(db, current_user.conjunto_id, payload)


# ── Conceptos ─────────────────────────────────────────────────────────────────

@router.get("/conceptos", response_model=list[schemas.ConceptoMovimientoOut])
def get_conceptos(
    tipo: str | None = Query(None, pattern="^(cargo|abono|ingreso|egreso)$"),
    todos: bool = Query(False),
    current_user=Depends(role_required(["admin"])),
    db: Session = Depends(get_db),
):
    return finanzas_service.list_conceptos(
        db, current_user.conjunto_id, tipo=tipo, solo_activos=not todos
    )


@router.post("/conceptos", response_model=schemas.ConceptoMovimientoOut, status_code=201)
def post_concepto(
    payload: schemas.ConceptoMovimientoCreate,
    current_user=Depends(role_required(["admin"])),
    db: Session = Depends(get_db),
):
    return finanzas_service.create_concepto(db, current_user.conjunto_id, payload)


@router.patch("/conceptos/{concepto_id}", response_model=schemas.ConceptoMovimientoOut)
def patch_concepto(
    concepto_id: int,
    payload: schemas.ConceptoMovimientoUpdate,
    current_user=Depends(role_required(["admin"])),
    db: Session = Depends(get_db),
):
    item = finanzas_service.update_concepto(db, current_user.conjunto_id, concepto_id, payload)
    if not item:
        raise HTTPException(status_code=404, detail="Concepto no encontrado")
    return item


# ── Cuotas ────────────────────────────────────────────────────────────────────

@router.post("/generar-cuotas", response_model=schemas.GenerarCuotasOut)
def post_generar_cuotas(
    payload: schemas.GenerarCuotasIn,
    current_user=Depends(role_required(["admin"])),
    db: Session = Depends(get_db),
):
    return finanzas_service.generar_cuotas(
        db, current_user.conjunto_id, payload.periodo, current_user.username
    )


# ── Cartera ───────────────────────────────────────────────────────────────────

@router.get("/cartera", response_model=list[schemas.CarteraItemOut])
def get_cartera(
    torre: str | None = None,
    estado: str | None = Query(None, pattern="^(al_dia|en_mora|todos)$"),
    saldo_min: int | None = None,
    saldo_max: int | None = None,
    current_user=Depends(role_required(["admin"])),
    db: Session = Depends(get_db),
):
    return finanzas_service.list_cartera(
        db,
        current_user.conjunto_id,
        torre=torre,
        estado=estado,
        saldo_min=saldo_min,
        saldo_max=saldo_max,
    )


# ── Estado de cuenta ──────────────────────────────────────────────────────────

@router.get("/propietarios/{uid}/estado-cuenta", response_model=schemas.EstadoCuentaOut)
def get_estado_cuenta(
    uid: str,
    current_user=Depends(role_required(["admin"])),
    db: Session = Depends(get_db),
):
    cuenta = finanzas_service.estado_cuenta(db, current_user.conjunto_id, uid)
    if not cuenta:
        raise HTTPException(status_code=404, detail="Propietario no encontrado")
    return cuenta


@router.post("/propietarios/{uid}/movimientos", response_model=schemas.EstadoCuentaOut)
def post_movimiento(
    uid: str,
    payload: schemas.MovimientoCarteraCreate,
    current_user=Depends(role_required(["admin"])),
    db: Session = Depends(get_db),
):
    try:
        cuenta = finanzas_service.crear_movimiento_cartera(
            db, current_user.conjunto_id, uid, payload, current_user.username
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if not cuenta:
        raise HTTPException(status_code=404, detail="Propietario no encontrado")
    return cuenta


@router.get("/propietarios/{uid}/estado-cuenta.pdf")
def export_estado_cuenta_pdf(
    uid: str,
    current_user=Depends(role_required(["admin"])),
    db: Session = Depends(get_db),
):
    cuenta = finanzas_service.estado_cuenta(db, current_user.conjunto_id, uid)
    if not cuenta:
        raise HTTPException(status_code=404, detail="Propietario no encontrado")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.6 * inch, bottomMargin=0.6 * inch)
    styles = getSampleStyleSheet()
    title = ParagraphStyle("TitleFin", parent=styles["Heading1"], fontSize=14, spaceAfter=12)
    story = [
        Paragraph("Estado de cuenta", title),
        Paragraph(
            f"<b>{cuenta.nombre}</b> — Torre {cuenta.torre} Apto {cuenta.apartamento} "
            f"(UID {cuenta.uid})",
            styles["Normal"],
        ),
        Paragraph(
            f"Saldo: {_centavos_a_cop(cuenta.saldo_centavos)} COP — Estado: {cuenta.estado_cuenta}",
            styles["Normal"],
        ),
        Spacer(1, 12),
    ]

    data = [["Fecha", "Tipo", "Concepto", "Monto", "Saldo"]]
    for m in cuenta.movimientos:
        data.append([
            m.fecha.isoformat(),
            m.tipo,
            m.concepto_nombre or m.referencia or "—",
            _centavos_a_cop(m.monto_centavos),
            _centavos_a_cop(m.saldo_acumulado_centavos),
        ])
    if len(data) == 1:
        data.append(["—", "—", "Sin movimientos", "—", "—"])

    table = Table(data, colWidths=[1.1 * inch, 0.8 * inch, 2.2 * inch, 1.1 * inch, 1.1 * inch])
    table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ])
    )
    story.append(table)
    story.append(Spacer(1, 10))
    story.append(
        Paragraph(
            f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            styles["Normal"],
        )
    )
    doc.build(story)
    buffer.seek(0)
    filename = f"estado-cuenta-{cuenta.torre}-{cuenta.apartamento}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/propietarios/{uid}/estado-cuenta.xlsx")
def export_estado_cuenta_xlsx(
    uid: str,
    current_user=Depends(role_required(["admin"])),
    db: Session = Depends(get_db),
):
    cuenta = finanzas_service.estado_cuenta(db, current_user.conjunto_id, uid)
    if not cuenta:
        raise HTTPException(status_code=404, detail="Propietario no encontrado")

    wb = Workbook()
    ws = wb.active
    ws.title = "Estado de cuenta"
    ws.append(["Nombre", cuenta.nombre])
    ws.append(["Torre", cuenta.torre])
    ws.append(["Apartamento", cuenta.apartamento])
    ws.append(["UID", cuenta.uid])
    ws.append(["Saldo COP", cuenta.saldo_centavos / 100])
    ws.append(["Estado", cuenta.estado_cuenta])
    ws.append([])
    ws.append(["Fecha", "Tipo", "Concepto", "Monto COP", "Saldo acumulado COP", "Referencia", "Notas"])
    for m in cuenta.movimientos:
        ws.append([
            m.fecha.isoformat(),
            m.tipo,
            m.concepto_nombre or "",
            m.monto_centavos / 100,
            m.saldo_acumulado_centavos / 100,
            m.referencia or "",
            m.notas or "",
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"estado-cuenta-{cuenta.torre}-{cuenta.apartamento}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Caja ──────────────────────────────────────────────────────────────────────

@router.get("/caja", response_model=list[schemas.MovimientoCajaOut])
def get_caja(
    periodo: str | None = Query(None, pattern=r"^\d{4}-\d{2}$"),
    tipo: str | None = Query(None, pattern="^(ingreso|egreso)$"),
    current_user=Depends(role_required(["admin"])),
    db: Session = Depends(get_db),
):
    return finanzas_service.list_caja(db, current_user.conjunto_id, periodo=periodo, tipo=tipo)


@router.post("/caja", response_model=schemas.MovimientoCajaOut, status_code=201)
def post_caja(
    payload: schemas.MovimientoCajaCreate,
    current_user=Depends(role_required(["admin"])),
    db: Session = Depends(get_db),
):
    try:
        return finanzas_service.crear_movimiento_caja(
            db, current_user.conjunto_id, payload, current_user.username
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


# ── Alertas ───────────────────────────────────────────────────────────────────

@router.get("/alertas", response_model=list[schemas.AlertaFinancieraOut])
def get_alertas(
    refrescar: bool = Query(True),
    solo_no_leidas: bool = Query(False),
    current_user=Depends(role_required(["admin"])),
    db: Session = Depends(get_db),
):
    if refrescar:
        return finanzas_service.refresh_alertas(db, current_user.conjunto_id)
    return finanzas_service.list_alertas(db, current_user.conjunto_id, solo_no_leidas)


@router.patch("/alertas/{alerta_id}/leer", response_model=schemas.AlertaFinancieraOut)
def patch_alerta_leer(
    alerta_id: int,
    current_user=Depends(role_required(["admin"])),
    db: Session = Depends(get_db),
):
    alerta = finanzas_service.marcar_alerta_leida(db, current_user.conjunto_id, alerta_id)
    if not alerta:
        raise HTTPException(status_code=404, detail="Alerta no encontrada")
    return alerta


# ── Recordatorio ──────────────────────────────────────────────────────────────

@router.post("/propietarios/{uid}/recordatorio", response_model=schemas.TelegramNotificationOut)
async def post_recordatorio(
    uid: str,
    current_user=Depends(role_required(["admin"])),
    db: Session = Depends(get_db),
):
    result = finanzas_service.mensaje_recordatorio(db, current_user.conjunto_id, uid)
    if not result:
        raise HTTPException(status_code=404, detail="Propietario no encontrado")
    propietario, mensaje = result
    if not propietario.telegram_chat_id:
        raise HTTPException(status_code=400, detail="El propietario no tiene Telegram vinculado")
    sent = await enviar_notificacion_telegram(db, propietario.id, mensaje)
    if not sent:
        raise HTTPException(status_code=502, detail="No se pudo enviar el recordatorio por Telegram")
    return schemas.TelegramNotificationOut(detail="Recordatorio enviado")


# ── Reportes ──────────────────────────────────────────────────────────────────

@router.get("/reportes/resumen.pdf")
def reporte_resumen_pdf(
    periodo: str | None = Query(None, pattern=r"^\d{4}-\d{2}$"),
    current_user=Depends(role_required(["admin"])),
    db: Session = Depends(get_db),
):
    cartera = finanzas_service.list_cartera(db, current_user.conjunto_id)
    caja = finanzas_service.list_caja(db, current_user.conjunto_id, periodo=periodo)

    total_cartera = sum(i.saldo_centavos for i in cartera if i.saldo_centavos > 0)
    ingresos = sum(m.monto_centavos for m in caja if m.tipo == "ingreso")
    egresos = sum(m.monto_centavos for m in caja if m.tipo == "egreso")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.6 * inch)
    styles = getSampleStyleSheet()
    story = [
        Paragraph("Reporte financiero", styles["Heading1"]),
        Paragraph(f"Periodo: {periodo or 'Todos'}", styles["Normal"]),
        Paragraph(f"Cartera vencida total: {_centavos_a_cop(total_cartera)} COP", styles["Normal"]),
        Paragraph(f"Ingresos caja: {_centavos_a_cop(ingresos)} COP", styles["Normal"]),
        Paragraph(f"Egresos caja: {_centavos_a_cop(egresos)} COP", styles["Normal"]),
        Spacer(1, 12),
        Paragraph("Cartera por apartamento", styles["Heading2"]),
    ]

    data = [["Torre", "Apto", "Nombre", "Saldo", "Estado"]]
    for item in cartera:
        data.append([
            item.torre,
            item.apartamento,
            item.nombre[:28],
            _centavos_a_cop(item.saldo_centavos),
            item.estado_cuenta,
        ])
    table = Table(data, colWidths=[0.7 * inch, 0.7 * inch, 2.4 * inch, 1.2 * inch, 1 * inch])
    table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ])
    )
    story.append(table)
    doc.build(story)
    buffer.seek(0)
    filename = f"reporte-financiero-{periodo or 'completo'}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/reportes/resumen.xlsx")
def reporte_resumen_xlsx(
    periodo: str | None = Query(None, pattern=r"^\d{4}-\d{2}$"),
    current_user=Depends(role_required(["admin"])),
    db: Session = Depends(get_db),
):
    cartera = finanzas_service.list_cartera(db, current_user.conjunto_id)
    caja = finanzas_service.list_caja(db, current_user.conjunto_id, periodo=periodo)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Cartera"
    ws1.append(["UID", "Nombre", "Torre", "Apartamento", "Saldo COP", "Estado", "Último pago", "Próximo vencimiento"])
    for item in cartera:
        ws1.append([
            item.uid,
            item.nombre,
            item.torre,
            item.apartamento,
            item.saldo_centavos / 100,
            item.estado_cuenta,
            item.ultimo_pago.isoformat() if item.ultimo_pago else "",
            item.proximo_vencimiento.isoformat() if item.proximo_vencimiento else "",
        ])

    ws2 = wb.create_sheet("Caja")
    ws2.append(["Fecha", "Tipo", "Concepto", "Monto COP", "Periodo", "Referencia", "Notas"])
    for m in caja:
        ws2.append([
            m.fecha.isoformat(),
            m.tipo,
            m.concepto_nombre or "",
            m.monto_centavos / 100,
            m.periodo or "",
            m.referencia or "",
            m.notas or "",
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"reporte-financiero-{periodo or 'completo'}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
